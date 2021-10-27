import os
from os.path import exists
from selenium import webdriver
from pathlib import Path
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import re
import time
import openpyxl
import PyPDF2
import csv

# =============================================================
# =================        class          =====================
# =============================================================

#Classe para configurar e iniciar o browser
class browserNavigate:

    browser = webdriver.Firefox
    geckodriver = Service(os.path.join(Path(os.path.dirname(os.path.abspath(__file__))), "Drivers", "geckodriver.exe"))

    def __init__(self):
        self.geckodriver = Service(os.path.join(Path(os.path.dirname(os.path.abspath(__file__))), "Drivers", "geckodriver.exe"))

    #Inicia o firefox
    def openFirefox(self, URL):

        options = Options()
        options.set_preference("browser.preferences.instantApply", True)
        options.set_preference("browser.helperApps.neverAsk.openFile", "application/pdf")
        options.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/pdf")
        options.set_preference('browser.helperApps.alwaysAsk.force', False)
        options.set_preference("browser.download.dir", os.path.join(Path(os.path.dirname(os.path.abspath(__file__))), 'Output'))
        options.set_preference("browser.download.folderList", 2)
        options.set_preference("browser.download.manager.showWhenStarting", False)
        options.set_preference("browser.download.manager.showAlertOnComplete", False)
        options.set_preference('browser.download.manager.alertOnEXEOpen', False)
        options.set_preference('browser.download.manager.focusWhenStarting', False)
        options.set_preference('browser.download.manager.useWindow', False)
        options.set_preference('browser.download.manager.showAlertOnComplete', False)
        options.set_preference('browser.download.manager.closeWhenDone', False)
        options.set_preference('pdfjs.disabled', True)
        self.browser = webdriver.Firefox(options= options, service=self.geckodriver)
        if URL != "":
            self.browser.get(URL)

#Classe main do processo, realizar Todoo fluxo end-to-end
class main():
    agenciaAnalisar = ""

    #Cria a pasta output na pasta do projeto
    if not exists(Path(os.path.join(Path(os.path.dirname(os.path.abspath(__file__)), 'Output')))):
        os.mkdir(Path(os.path.join(Path(os.path.dirname(os.path.abspath(__file__)), 'Output'))))

    #Abre o navegador
    browserNavigate = browserNavigate()
    browserNavigate.openFirefox("https://itdashboard.gov/")

    #Clica em Dive In
    element = WebDriverWait(browserNavigate.browser, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="node-23"]/div/div/div/div/div/div/div/a')))
    element.click()

    #Espera a pagina carregar
    WebDriverWait(browserNavigate.browser, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="agency-tiles-widget"]/div/div[1]/div[1]/div/div/div/div[1]/a/span[1]')))

    #Cria a planilha excel dentro da pasta output na pasta do projeto
    workbook = openpyxl.Workbook()
    workbook.save(Path(os.path.join(Path(os.path.dirname(os.path.abspath(__file__)),'Output','Agencias.xlsx'))))
    workbook = openpyxl.load_workbook(Path(os.path.join(Path(os.path.dirname(os.path.abspath(__file__)),'Output','Agencias.xlsx'))))
    sheet = workbook['Sheet']
    sheet.title = 'Agencias'

    sheet['A1'] = 'Agências'
    sheet['B1'] = 'Montante'

    row = 1
    j=0
    fimPagina = False

    #Extrai todas as agencias e montantes na pagina
    while not fimPagina:
        j += 1
        try:
            for i in range(3):
                row += 1
                element = browserNavigate.browser.find_element(By.XPATH, '//*[@id="agency-tiles-widget"]/div/div['+ str(j) +']/div['+ str(i+1) +']/div/div/div/div[1]/a/span[1]')
                agencia = element.text
                sheet['A'+str(row)] = agencia

                element = browserNavigate.browser.find_element(By.XPATH, '//*[@id="agency-tiles-widget"]/div/div['+ str(j) +']/div['+ str(i+1) +']/div/div/div/div[1]/a/span[2]')
                montante = element.text
                sheet['B' + str(row)] = montante
        except:
            fimPagina = True

    #Lê o arquivo de configuração e obtem o nome da agencia que irá extrair os dados e arquivos
    with open((Path(os.path.join(Path(os.path.dirname(os.path.abspath(__file__)),'Config','Config.csv')))), 'r') as file:
        reader = csv.reader(file)
        for row in reader:
            agenciaAnalisar = row[1]
    try:
        xpath = '//*[text()="' + agenciaAnalisar + '"]'
        element = browserNavigate.browser.find_element(By.XPATH, xpath)
        print(element.text)
        element.click()
    except:
        print("Agência Não Encontrada")
        exit()

    #Espera a tabela carregar
    WebDriverWait(browserNavigate.browser, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="investments-table-object"]/tbody/tr[1]/td[1]')))

    #Seleciona todos as entradas, para evitar paginação o que iria ser menos otimizado no processo
    element = Select(browserNavigate.browser.find_element(By.XPATH, '//*[@id="investments-table-object_length"]/label/select'))
    element.select_by_index(3)

    #Espera todas as entradas serem carregadas
    element = browserNavigate.browser.find_element(By.XPATH, '//*[@id="investments-table-object_info"]')
    entries = element.text
    result1 = re.search('to(.*)of', entries)
    result2 = re.search('of(.*)entr', entries)
    while result1.group(1).strip() != result2.group(1).strip():
        element = browserNavigate.browser.find_element(By.XPATH, '//*[@id="investments-table-object_info"]')
        entries = element.text
        result1 = re.search('to(.*)of', entries)
        result2 = re.search('of(.*)entr', entries)

    #Percorre todas as linhas da tabela
    time.sleep(2)
    lookPropertyRow = '//*[@id="investments-table-object"]/tbody/tr/td[1]'
    lookPropertyColumn = '//*[@id="investments-table-object"]/tbody/tr[1]/td'
    lookPropertyCell = '//*[@id="investments-table-object"]/tbody/tr[rowID]/td[columnID]'

    row = browserNavigate.browser.find_elements(By.XPATH, lookPropertyRow)
    column = browserNavigate.browser.find_elements(By.XPATH, lookPropertyColumn)

    i = 1
    table = []
    #Para cada linha
    if len(column) > 0:
        while (i <= len(row)):
            aux = []
            j = 1
            while (j <= len(column)):
                cell = lookPropertyCell.replace("rowID", str(i))
                cell = cell.replace("columnID", str(j))
                try:
                    item = browserNavigate.browser.find_element(By.XPATH, cell)
                    aux.append(item.text.lstrip().rstrip())
                    #Verifica se a coluna UII tem link, caso tenha, faz download e analisa o PDF (fiz somente a extração dos dados), não fiz a comparação
                    if j == 1:
                        link = browserNavigate.browser.find_element(By.XPATH, cell + "/a").get_attribute('href')
                        if link != None:
                            fileName = browserNavigate.browser.find_element(By.XPATH, cell).text + ".pdf"
                            browserNavigate.browser.execute_script("window.open('"+link+"'),'_blank'")
                            handle = browserNavigate.browser.window_handles[1]
                            browserNavigate.browser.switch_to.window(handle)
                            WebDriverWait(browserNavigate.browser, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="business-case-pdf"]/a')))
                            element = browserNavigate.browser.find_element(By.XPATH, '//*[@id="business-case-pdf"]/a')
                            element.click()
                            file_exists = exists(os.path.join(Path(os.path.dirname(os.path.abspath(__file__))), 'Output', fileName))
                            while not file_exists:
                                file_exists = exists(os.path.join(Path(os.path.dirname(os.path.abspath(__file__))), 'Output', fileName))
                            browserNavigate.browser.close()

                            file = open(os.path.join(Path(os.path.dirname(os.path.abspath(__file__))), 'Output', fileName), 'rb')
                            fileReader = PyPDF2.PdfFileReader(file)
                            text = fileReader.getPage(0).extractText()
                            nameOfInvestment = re.search('(?<=Name\sof\sthis\sInvestment:\\n\s\\n).*(?=\\n2\.\sUnique)',text).group(0)
                            investmentIdentifier = re.search('(?<=Unique\sInvestment\sIdentifier\s\(UII\):\\n\s\\n).*(?=\\nSection)',text).group(0)

                            handle = browserNavigate.browser.window_handles[0]
                            browserNavigate.browser.switch_to.window(handle)
                except:
                    aux.append("")
                j += 1
            i += 1
            table.append(aux)

    #Após a tabela extraida e todos os arquivos baixados, insere os dados em uma nova sheet no arquivo excel
    workbook.create_sheet('Individual Investments')
    sheet = workbook['Individual Investments']
    sheet['A1'] = browserNavigate.browser.find_element(By.XPATH, '//*[@id="investments-table-object_wrapper"]/div[3]/div[1]/div/table/thead/tr[2]/th[1]').text
    sheet['B1'] = browserNavigate.browser.find_element(By.XPATH, '//*[@id="investments-table-object_wrapper"]/div[3]/div[1]/div/table/thead/tr[2]/th[2]').text
    sheet['C1'] = browserNavigate.browser.find_element(By.XPATH, '//*[@id="investments-table-object_wrapper"]/div[3]/div[1]/div/table/thead/tr[2]/th[3]').text
    sheet['D1'] = browserNavigate.browser.find_element(By.XPATH, '//*[@id="investments-table-object_wrapper"]/div[3]/div[1]/div/table/thead/tr[2]/th[4]').text
    sheet['E1'] = browserNavigate.browser.find_element(By.XPATH, '//*[@id="investments-table-object_wrapper"]/div[3]/div[1]/div/table/thead/tr[2]/th[5]').text
    sheet['F1'] = browserNavigate.browser.find_element(By.XPATH, '//*[@id="investments-table-object_wrapper"]/div[3]/div[1]/div/table/thead/tr[2]/th[6]').text
    sheet['G1'] = browserNavigate.browser.find_element(By.XPATH, '//*[@id="investments-table-object_wrapper"]/div[3]/div[1]/div/table/thead/tr[2]/th[7]').text

    rowsCount = len(table)

    for i in range(rowsCount):
        sheet['A' + str(i+2)] = table[i][0]
        sheet['B' + str(i+2)] = table[i][1]
        sheet['C' + str(i+2)] = table[i][2]
        sheet['D' + str(i+2)] = table[i][3]
        sheet['E' + str(i+2)] = table[i][4]
        sheet['F' + str(i+2)] = table[i][5]
        sheet['G' + str(i+2)] = table[i][6]
    workbook.save(Path(os.path.join(Path(os.path.dirname(os.path.abspath(__file__)), 'Output', 'Agencias.xlsx'))))
    browserNavigate.browser.close()


    #FIM#


